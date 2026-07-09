from collections import defaultdict
from typing import Optional
import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment, numbers

header = NamedStyle(name="header")
header.font = Font(bold=True)
header.fill = PatternFill(
                        "solid",
                        fgColor="BDBDBD",
                        #bgColor="BDBDBD",
                        )

alternating = NamedStyle(name="alternating")
alternating.fill = PatternFill(
                        "solid",
                        fgColor="F3F3F3",
                        #bgColor="BDBDBD",
                        )

#https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
blink_class_format = {
    0:
    {#Node1,Node2,TotUniq,Perc,Infores,Predicates
        "row_widths":[191/7, 162/7, 106/7, 91/7, 261/7, 624/7],
        "number_format":[None,None,numbers.FORMAT_GENERAL,numbers.FORMAT_NUMBER_00,None,None]
    }
}
formats = {
    "biolink_summary_format":blink_class_format
}

def formatInforesSummary(wb:openpyxl.Workbook):
    spoq_col_lens = {
        "default":16.57,
        "Knowledge-Level Terms":18.57,
        "Agent-Type Terms": 18.57,
        "Publication Counts": 32.29,
        "Evidence Counts": 32.29,
        "Edge Properties":43.86,
        "original_json":79.4
    }
    spoq_col_lens = defaultdict(lambda:spoq_col_lens["default"],spoq_col_lens)
    for ws_idx,ws in enumerate(wb.worksheets):
        for col_idx,column_cells in enumerate(ws.columns):
            header_cell = column_cells[0]
            header_text = header_cell.value
            col_len = spoq_col_lens[header_text]
            ws.column_dimensions[header_cell.column_letter].width = col_len
            #print(ws_idx,col_idx, head_col.value,x)

    hide_list = ["SCat (Actual)","OCat (Actual)","SPQO Tuple"]
    for ws_idx,ws in enumerate(wb.worksheets):
        for col_idx,column_cells in enumerate(ws.columns):
            header_cell = column_cells[0]
            header_text = header_cell.value
            if(header_text in hide_list):
                ws.column_dimensions[header_cell.column_letter].hidden = True
            #print(ws_idx,col_idx, head_col.value,x)

def formatXlsx(xlsx_file,output_xlsx_file,format:Optional[str]=None,infores_summary_format:bool=False):
    format_style=None
    if(format!=None):format_style = formats[format]
    wb = openpyxl.load_workbook(filename = xlsx_file)
    try:
        wb.add_named_style(header)
        wb.add_named_style(alternating)
    except ValueError:
        pass
        #Bolded header, alternating colors
    for ws in wb.worksheets:
        for i,row_cells in enumerate(ws.rows):
            if(i==0):
                for cell in row_cells:
                    cell.style = 'header'

            elif(i%2==0):
                for cell in row_cells:
                    cell.style='alternating'

    #Columns width of title
    for ws in wb.worksheets:
        for i,column_cells in enumerate(ws.columns):
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length # type: ignore

    for ws in wb.worksheets:
        for row in ws.rows:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
    
    if(format_style!=None):
        #Format widths if available
        
        for ws_idx, ws in enumerate(wb.worksheets):
            if(ws_idx not in format_style): continue
            for col_idx,column_cells in enumerate(ws.columns):
                length = format_style[ws_idx]["row_widths"][col_idx]
                ws.column_dimensions[column_cells[0].column_letter].width = length # type: ignore
        
        #Format number style if available.
        for ws_idx, ws in enumerate(wb.worksheets):
            if(ws_idx not in format_style): continue
            for col_idx,column_cells in enumerate(ws.columns):
                if(format_style[ws_idx]["number_format"][col_idx]==None):continue
                col_format = format_style[ws_idx]["number_format"][col_idx]
#                print(col_idx,col_format)
                for cell in column_cells:
                    cell.number_format = col_format
    if(infores_summary_format):
        formatInforesSummary(wb)
    
    wb.save(output_xlsx_file)



if(__name__=="__main__"):
    import sys
    from pathlib import Path
    import os
    xlsx_file = sys.argv[1]
    p = Path(xlsx_file)
    formatted_dir = os.path.join("FORMATTED",p.parent)
    os.makedirs(formatted_dir,exist_ok=True)
    out_xlsx_file = os.path.join(formatted_dir,p.name)
    if((len(sys.argv))==3):
        format = sys.argv[2]
    else:
        format = None
    print(f"Fromatting {xlsx_file}")
    try: 
        formatXlsx(xlsx_file,out_xlsx_file,format,True)
    except Exception as e:
        print(f"!!!Could not format {xlsx_file} --- Got error {str(e)}!!!")
    print(f"Fromatted {xlsx_file} --- wrote to {out_xlsx_file}")